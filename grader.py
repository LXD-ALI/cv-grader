#!/usr/bin/python3
"""
Excel Worksheet Grader - Local Version

This module provides the grading functionality for local testing.
"""

import openpyxl

# Worksheet names
STUDENT_SHEET_NAME = "blank"
SOLUTION_SHEET_NAME = "solution"

def grade_excel_worksheet(student_file_path, solution_file_path="solution.xlsx"):
    """
    Grade Excel worksheet by comparing Y/N values in row 1
    
    Args:
        student_file_path: Path to the student's Excel file
        solution_file_path: Path to the solution Excel file
        
    Returns:
        Dictionary with score and feedback
    """
    try:
        # Load workbooks
        student_wb = openpyxl.load_workbook(student_file_path, data_only=True)
        solution_wb = openpyxl.load_workbook(solution_file_path, data_only=True)
        
        # Verify sheets exist
        if STUDENT_SHEET_NAME not in student_wb.sheetnames:
            return {
                "score": 0.0,
                "feedback": f"Error: Worksheet '{STUDENT_SHEET_NAME}' not found in your submission."
            }
        
        if SOLUTION_SHEET_NAME not in solution_wb.sheetnames:
            return {
                "score": 0.0,
                "feedback": f"Error: Solution worksheet not found."
            }
        
        # Get sheets
        student_sheet = student_wb[STUDENT_SHEET_NAME]
        solution_sheet = solution_wb[SOLUTION_SHEET_NAME]
        
        # Get Y/N values from row 1 in both sheets
        matches = 0
        total_cells = 0
        max_col = max(solution_sheet.max_column, student_sheet.max_column)
        
        # Start from column E (index 5 in openpyxl)
        for col_idx in range(5, max_col + 1):
            student_cell = student_sheet.cell(row=1, column=col_idx)
            solution_cell = solution_sheet.cell(row=1, column=col_idx)
            
            if student_cell.value is not None and solution_cell.value is not None:
                total_cells += 1
                if student_cell.value == solution_cell.value:
                    matches += 1
        
        # Calculate score and generate feedback
        score = matches / total_cells if total_cells > 0 else 0.0
        percentage = score * 100
        feedback = f"Your score: {percentage:.2f}%\nYou correctly matched {matches} out of {total_cells} cells."
        
        return {
            "score": score,
            "feedback": feedback,
            "matches": matches,
            "total_cells": total_cells
        }
        
    except Exception as e:
        return {
            "score": 0.0,
            "feedback": f"Error grading your submission: {str(e)}"
        }
